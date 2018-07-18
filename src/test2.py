import requests
import re
import bs4
import time
import os
import openpyxl
class spider(object):
    def __init__(self,url):
        super(spider,self).__init__()
        self.url = url
        # self.word = word

    def getHtml(self,word):
        kv = {'user-agent':'Mazilia/5.0'}  # 模拟浏览器访问
        kw = {'wd':word}
        try:
            r = requests.get(self.url,headers=kv, params=kw)
            r.raise_for_status()
            r.encoding = r.apparent_encoding
            return r.text
        except:
            return "error"
    

    def conductHtml(self,html):
        i = 0
        soup = bs4.BeautifulSoup(html,'html.parser')
        tag = soup.find('div',id='content_left')
        # print(tag)
        tag_span = tag.find_all('span')
        # print(tag_span)
        # print('网站：')
        print('-'*20)
        value = []
        for span_ in tag_span:
            # print(span_.string)
            if span_.string =='广告':
                div_a = span_.parent.parent
                # print(div_a)
                i+=1
               # 获取标题 
                div_title = span_.parent.parent.parent
                row_value = []
                for sibling_title in div_title.previous_siblings:

                    try:
                        a_title = sibling_title.h3.a
                        if a_title.string:
                            # print("Title:",a_title.string)
                            row_value.append(a_title.string)
                        else:
                            a_title = str(a_title)
                            patten = re.compile(r'\<.*?\>')
                            title = patten.sub('',a_title)
                            # print(a_title)
                            # print("Title:",title)
                            row_value.append(title)
                    except:
                        pass


                # 获取网址
                for sibling_a in div_a.previous_siblings:
                    # print(sibling_a)
                    try:
                        # print('-'*10)
                        # print("Address:",sibling_a.span.string)
                        row_value.append(sibling_a.span.string)
                    except:
                        pass
                
                # print('-'*20)
                value.append(row_value)
        # print('数量：',i)
        return i,value
        # print(tag)

    def re_exe(self,cmd, inc):
        while True:
            time.sleep(inc)
            os.system(cmd)
    
    def readExcel(self,path,sheetName):
        wb = openpyxl.load_workbook(path)
        sheet = wb.get_sheet_by_name(sheetName)
        max_row = sheet.max_row
        return max_row

    def writeExcel(self,path,sheetName,max_row,value_,mount):
        wb = openpyxl.load_workbook(path)
        sheet = wb.get_sheet_by_name(sheetName)
        for i in range(0,len(value_)):
            for j in range(0, len(value_[i])):
                sheet.cell(row = i+1+max_row, column = j+1, value = str(value_[i][j]))
        sheet.cell(row = len(value_)+1+max_row,column = 1, value = str(mount)) 
        wb.save(path)


if __name__ == "__main__":
    url = "http://www.baidu.com/s"
    word = "手机"
    sp = spider(url)
    htmlText = sp.getHtml(word)
    num,value_ = sp.conductHtml(htmlText)

    max_row = sp.readExcel('test2.xlsx','Sheet1')
    sp.writeExcel('test2.xlsx','Sheet1',max_row, value_,num)
    # 重复执行： inc：为执行时间，单位秒
    cmd = "python3 test2.py"
    inc = 20
    sp.re_exe(cmd,inc)


