import openpyxl

def write07Excel(path,max_row):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('2007测试表')
    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1+max_row, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")

def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('2007测试表')
    max_row = sheet.max_row

    return max_row



max_row = read07Excel('test.xlsx')
print(max_row)
write07Excel('test.xlsx',max_row)